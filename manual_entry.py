#!/usr/bin/env python3
"""
Manual Trading Data Entry Tool
Use this if OCR extraction is incomplete
"""

import sys
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from trading_data_analyzer import TradingDataExtractor

def manual_entry():
    """Allow manual entry of trading data"""
    trading_data = []
    
    print("\n" + "=" * 60)
    print("MANUAL TRADING DATA ENTRY")
    print("=" * 60)
    print("\nEnter your trading data (press Enter with no input when done)\n")
    
    while True:
        print(f"\nTrade {len(trading_data) + 1}:")
        
        # Get direction
        direction = input("  Direction (BUY/SELL): ").strip().upper()
        if not direction:
            if trading_data:
                break
            else:
                print("  Please enter BUY or SELL")
                continue
        
        if direction not in ['BUY', 'SELL']:
            print("  Invalid direction. Enter BUY or SELL")
            continue
        
        # Get quantity
        try:
            quantity = int(input("  Quantity: ").strip())
            if quantity <= 0:
                print("  Quantity must be positive")
                continue
        except ValueError:
            print("  Invalid quantity. Enter a number")
            continue
        
        # Get rate
        try:
            rate = float(input("  Market Rate: ").strip())
            if rate <= 0:
                print("  Rate must be positive")
                continue
        except ValueError:
            print("  Invalid rate. Enter a number")
            continue
        
        trading_data.append({
            'direction': direction,
            'quantity': quantity,
            'rate': rate
        })
        print(f"  ✓ Trade added: {direction} {quantity} @ {rate}")
    
    return trading_data

def generate_excel_from_data(trading_data, output_path='trading_analysis_manual.xlsx'):
    """Generate Excel file from manually entered data"""
    if not trading_data:
        print("No data to process")
        return
    
    print("\n" + "=" * 60)
    print("GENERATING EXCEL REPORT")
    print("=" * 60)
    
    # Separate BUY and SELL
    buy_trades = [t for t in trading_data if t['direction'] == 'BUY']
    sell_trades = [t for t in trading_data if t['direction'] == 'SELL']
    
    # Calculate values
    for trade in trading_data:
        trade['calculated'] = (trade['quantity'] / 250) * trade['rate']
    
    buy_total = sum(t['calculated'] for t in buy_trades)
    sell_total = sum(t['calculated'] for t in sell_trades)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trading Data"
    
    # Add title
    ws['A1'] = "TRADING DATA ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Add headers
    ws['A3'] = "Direction"
    ws['B3'] = "Quantity"
    ws['C3'] = "Market Rate"
    ws['D3'] = "Calculated Value (Qty/250 × Rate)"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}3']
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # Add data
    row = 4
    for trade in trading_data:
        ws[f'A{row}'] = trade['direction']
        ws[f'B{row}'] = trade['quantity']
        ws[f'C{row}'] = trade['rate']
        ws[f'D{row}'] = trade['calculated']
        
        ws[f'C{row}'].number_format = '0.0000'
        ws[f'D{row}'].number_format = '0.0000'
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = thin_border
        
        row += 1
    
    # Add summary
    summary_row = row + 2
    ws[f'A{summary_row}'] = "SUMMARY"
    ws[f'A{summary_row}'].font = Font(bold=True, size=12)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Type"
    ws[f'B{summary_row}'] = "Total"
    
    for col in ['A', 'B']:
        ws[f'{col}{summary_row}'].font = Font(bold=True)
        ws[f'{col}{summary_row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws[f'{col}{summary_row}'].border = thin_border
    
    summary_row += 1
    ws[f'A{summary_row}'] = "BUY"
    ws[f'B{summary_row}'] = buy_total
    ws[f'B{summary_row}'].number_format = '0.0000'
    ws[f'B{summary_row}'].font = Font(bold=True, color="008000")
    ws[f'B{summary_row}'].border = thin_border
    
    summary_row += 1
    ws[f'A{summary_row}'] = "SELL"
    ws[f'B{summary_row}'] = sell_total
    ws[f'B{summary_row}'].number_format = '0.0000'
    ws[f'B{summary_row}'].font = Font(bold=True, color="FF0000")
    ws[f'B{summary_row}'].border = thin_border
    
    summary_row += 1
    ws[f'A{summary_row}'] = "NET (BUY - SELL)"
    ws[f'B{summary_row}'] = buy_total - sell_total
    ws[f'B{summary_row}'].number_format = '0.0000'
    ws[f'B{summary_row}'].font = Font(bold=True)
    ws[f'B{summary_row}'].border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    
    # Save
    wb.save(output_path)
    print(f"\n✓ Excel report generated: {output_path}")
    
    # Print summary
    print("\n" + "=" * 60)
    print("ANALYSIS SUMMARY")
    print("=" * 60)
    print(f"BUY Count:  {len(buy_trades)}")
    print(f"SELL Count: {len(sell_trades)}")
    print(f"BUY Total:  {buy_total:,.4f}")
    print(f"SELL Total: {sell_total:,.4f}")
    print(f"NET (BUY - SELL): {buy_total - sell_total:,.4f}")
    print("=" * 60)

if __name__ == "__main__":
    trading_data = manual_entry()
    
    if trading_data:
        generate_excel_from_data(trading_data)
        print("\n✓ Done!")
    else:
        print("No data entered.")
