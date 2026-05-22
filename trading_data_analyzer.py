#!/usr/bin/env python3
"""
Trading Data Analysis Tool
Extracts trading data from images using OCR, performs calculations, and generates Excel reports.

Requirements:
pip install pytesseract pillow openpyxl opencv-python

Also requires Tesseract OCR to be installed:
- Windows: Download installer from https://github.com/UB-Mannheim/tesseract/wiki
- macOS: brew install tesseract
- Linux: sudo apt-get install tesseract-ocr
"""

import os
import sys
import re
import time
from pathlib import Path
from PIL import Image
import cv2
import pytesseract
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

MAX_OCR_IMAGE_DIMENSION = int(os.environ.get("MAX_OCR_IMAGE_DIMENSION", "3200"))
OCR_TIMEOUT_SECONDS = int(os.environ.get("OCR_TIMEOUT_SECONDS", "180"))
# --psm 6 = uniform block of text (default in old code).
# --psm 4 = single column of text of variable sizes (works well for trade tables).
TESSERACT_CONFIG = os.environ.get("TESSERACT_CONFIG", r"--oem 3 --psm 6")
TESSERACT_CONFIG_FALLBACK = os.environ.get("TESSERACT_CONFIG_FALLBACK", r"--oem 3 --psm 4")
MIN_WORD_CONFIDENCE = int(os.environ.get("MIN_WORD_CONFIDENCE", "30"))

# Configure Tesseract path for Windows
if sys.platform == 'win32':
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'


class StepLog:
    """Collects per-step records for a single image and optionally echoes them to stdout.

    Each entry is a dict with keys: step, status (ok|warn|error|info), message, detail, elapsed_ms.
    """

    def __init__(self, echo=True, prefix=""):
        self.entries = []
        self.echo = echo
        self.prefix = prefix
        self._start = time.perf_counter()

    def add(self, step, status, message, detail=""):
        elapsed_ms = int((time.perf_counter() - self._start) * 1000)
        entry = {
            "step": step,
            "status": status,
            "message": message,
            "detail": detail,
            "elapsed_ms": elapsed_ms,
        }
        self.entries.append(entry)
        if self.echo:
            icon = {"ok": "[ok]", "warn": "[!]", "error": "[x]", "info": "[.]"}.get(status, "[.]")
            try:
                print(f"  {icon} [{step}] {self.prefix}{message} ({elapsed_ms} ms)", flush=True)
                if detail:
                    snippet = detail if len(detail) < 400 else detail[:400] + "..."
                    for line in snippet.splitlines():
                        print(f"      {line}", flush=True)
            except UnicodeEncodeError:
                # Fall back to ASCII-only output on legacy terminals.
                safe_msg = message.encode("ascii", "replace").decode("ascii")
                print(f"  {icon} [{step}] {self.prefix}{safe_msg} ({elapsed_ms} ms)", flush=True)
        return entry

    def reset_timer(self):
        self._start = time.perf_counter()


def _strip_number_token(token):
    """Strip currency, parentheses, percent etc., return cleaned numeric string or None."""
    if token is None:
        return None
    t = token.strip().replace(",", "")
    t = t.strip("()[]{}$₹€£%")
    if t.endswith("."):
        t = t[:-1]
    if t.startswith("."):
        t = "0" + t
    if not t:
        return None
    if not re.fullmatch(r"-?\d+(\.\d+)?", t):
        return None
    return t


def _looks_like_time_or_date(token):
    return bool(re.search(r"\d:\d", token)) or bool(re.search(r"\d[/-]\d", token))


class TradingDataExtractor:
    def __init__(self):
        self.trading_data = []
        self.documents = {}
        # per-document steps log: doc_name -> list[dict]
        self.steps_by_doc = {}
        
    def preprocess_image(self, image_path, log=None):
        """Load image, downscale if huge, grayscale + Otsu binarize for OCR."""
        try:
            img = cv2.imread(str(image_path))
            if img is None:
                msg = f"Could not read image: {image_path}"
                if log: log.add("preprocess", "error", msg,
                                "Verify path (spaces/parentheses) and that the format is JPG/PNG/JPEG/TIFF.")
                else:
                    print(f"  ✗ {msg}")
                raise ValueError(msg)

            height, width = img.shape[:2]
            orig_size = f"{width}x{height}"
            largest_dimension = max(width, height)
            if largest_dimension > MAX_OCR_IMAGE_DIMENSION:
                scale = MAX_OCR_IMAGE_DIMENSION / largest_dimension
                new_size = (int(width * scale), int(height * scale))
                img = cv2.resize(img, new_size, interpolation=cv2.INTER_AREA)
                if log: log.add("preprocess", "info",
                                f"Resized {orig_size} -> {new_size[0]}x{new_size[1]} (cap={MAX_OCR_IMAGE_DIMENSION})")
            else:
                if log: log.add("preprocess", "info", f"Loaded image {orig_size}")

            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            # Otsu binarization tends to be both faster for Tesseract and more accurate
            # for clean screenshots; if the image is photo-like it might overdo it, so we
            # keep the grayscale around in case binarization fails.
            try:
                _, binarized = cv2.threshold(gray, 0, 255,
                                             cv2.THRESH_BINARY + cv2.THRESH_OTSU)
                if log: log.add("preprocess", "ok", "Grayscale + Otsu binarization applied")
                return binarized
            except cv2.error as exc:
                if log: log.add("preprocess", "warn", f"Binarization failed, using grayscale: {exc}")
                return gray
        except Exception as e:
            if log: log.add("preprocess", "error", f"Preprocess failed: {e}")
            else: print(f"Error preprocessing image {image_path}: {e}")
            return None

    def _run_tesseract_data(self, pil_img, config, log=None):
        """Run image_to_data and return list of word dicts with bboxes + confidence."""
        data = pytesseract.image_to_data(
            pil_img,
            config=config,
            timeout=OCR_TIMEOUT_SECONDS,
            output_type=pytesseract.Output.DICT,
        )
        words = []
        n = len(data.get("text", []))
        for i in range(n):
            text = (data["text"][i] or "").strip()
            if not text:
                continue
            try:
                conf = int(float(data["conf"][i]))
            except (TypeError, ValueError):
                conf = -1
            words.append({
                "text": text,
                "conf": conf,
                "block": data["block_num"][i],
                "par": data["par_num"][i],
                "line": data["line_num"][i],
                "word": data["word_num"][i],
                "left": data["left"][i],
                "top": data["top"][i],
                "width": data["width"][i],
                "height": data["height"][i],
            })
        if log:
            avg_conf = (sum(w["conf"] for w in words if w["conf"] >= 0) / max(1, sum(1 for w in words if w["conf"] >= 0))) if words else 0
            log.add("ocr", "ok",
                    f"OCR returned {len(words)} words (avg conf {avg_conf:.0f}, config '{config}')")
        return words

    def run_ocr(self, image_path, log=None):
        """Step 2: run OCR -> returns (raw_text, words). Falls back to a 2nd PSM if first is empty."""
        processed = self.preprocess_image(image_path, log=log)
        if processed is None:
            return None, []

        pil_img = Image.fromarray(processed)

        try:
            words = self._run_tesseract_data(pil_img, TESSERACT_CONFIG, log=log)
            if not words and TESSERACT_CONFIG_FALLBACK != TESSERACT_CONFIG:
                if log: log.add("ocr", "warn", "Primary OCR found 0 words, retrying with fallback PSM")
                words = self._run_tesseract_data(pil_img, TESSERACT_CONFIG_FALLBACK, log=log)
        except FileNotFoundError:
            msg = ("Tesseract not found. Install it: "
                   "Windows https://github.com/UB-Mannheim/tesseract/wiki, "
                   "macOS `brew install tesseract`, Linux `apt-get install tesseract-ocr`.")
            if log: log.add("ocr", "error", msg)
            else: print(msg)
            return None, []
        except Exception as e:
            if log: log.add("ocr", "error", f"OCR failed: {e}")
            else: print(f"Error extracting text from {image_path}: {e}")
            return None, []

        raw_text = self._reconstruct_text(words)
        if log:
            log.add("ocr", "info",
                    f"Reconstructed text: {len(raw_text)} chars, {len(raw_text.splitlines())} lines",
                    detail=raw_text[:600])
        return raw_text, words

    def _reconstruct_text(self, words):
        """Group words back into lines using Tesseract's (block, par, line) grouping."""
        if not words:
            return ""
        groups = {}
        for w in words:
            key = (w["block"], w["par"], w["line"])
            groups.setdefault(key, []).append(w)
        lines = []
        for key in sorted(groups.keys()):
            row = sorted(groups[key], key=lambda x: x["left"])
            lines.append(" ".join(w["text"] for w in row))
        return "\n".join(lines)

    # Kept for backward compat with code that imports it.
    def extract_text_from_image(self, image_path):
        text, _ = self.run_ocr(image_path, log=None)
        return text

    # ------------------------------------------------------------------ tabular

    def build_table_rows(self, words, log=None):
        """Step 3: group OCR words into tabular rows using Tesseract's own line grouping
        (block_num, par_num, line_num). Returns list of rows; each row is a list of word dicts
        sorted left-to-right."""
        if not words:
            if log: log.add("table", "warn", "No OCR words to assemble into rows")
            return []

        # Discard low-confidence noise words but keep words even with conf == -1
        # (Tesseract emits -1 for some merged tokens).
        filtered = [w for w in words if w["conf"] < 0 or w["conf"] >= MIN_WORD_CONFIDENCE]
        dropped = len(words) - len(filtered)

        groups = {}
        for w in filtered:
            key = (w["block"], w["par"], w["line"])
            groups.setdefault(key, []).append(w)

        rows = []
        for key in sorted(groups.keys()):
            row = sorted(groups[key], key=lambda x: x["left"])
            rows.append(row)

        if log:
            log.add("table", "ok",
                    f"Built {len(rows)} rows from {len(filtered)} words"
                    + (f" (dropped {dropped} low-conf)" if dropped else ""))
        return rows

    def parse_trades_from_rows(self, rows, log=None):
        """Step 4 (preferred): find BUY/SELL rows in tabular output and extract qty + rate.

        Positional heuristic (mirrors how a trade row is actually structured):
          - Skip tokens that look like times (12:34) or dates (12/05).
          - For each row containing a BUY or SELL token, walk the words left-to-right and
            collect (index, value, has_decimal) for every numeric.
          - Rate detection (column layout aware):
              * If the row has >=2 decimal numerics, the trade-report layout is usually
                `... Qty | Market Rate | Amount` where Amount = Qty * Market Rate. We test
                this by walking integers to the left of the second-to-last decimal and
                checking whether `int_value * second_to_last_decimal ~= last_decimal`
                (2% tolerance). If verified, Rate = second-to-last decimal, Qty = that
                integer.
              * If only one decimal exists, Rate = that decimal.
              * Otherwise Rate = right-most numeric (legacy fallback).
          - Quantity = the nearest INTEGER numeric immediately to the left of the rate.
            Fallback: the right-most numeric to the left of the rate, regardless of type.
        This avoids picking an Amount column or an order-ID as the rate / quantity.
        """
        trades = []
        considered = 0
        skipped_reasons = []
        for row_idx, row in enumerate(rows):
            tokens = [w["text"] for w in row]
            joined = " ".join(tokens)
            upper = joined.upper()
            if "BUY" not in upper and "SELL" not in upper:
                continue
            considered += 1
            direction = "BUY" if "BUY" in upper else "SELL"

            numerics = []  # list of (token_index, cleaned_str, float_value, has_decimal)
            for i, tok in enumerate(tokens):
                if _looks_like_time_or_date(tok):
                    continue
                cleaned = _strip_number_token(tok)
                if cleaned is None:
                    continue
                try:
                    value = float(cleaned)
                except ValueError:
                    continue
                numerics.append((i, cleaned, value, "." in cleaned))

            if len(numerics) < 2:
                skipped_reasons.append(f"row {row_idx}: only {len(numerics)} numeric token(s)")
                continue

            decimals = [e for e in numerics if e[3]]

            # ---- pick rate (and possibly quantity together, when we can verify Qty*Rate=Amount)
            rate_entry = None
            quantity_entry = None

            if len(decimals) >= 2:
                # Try to match the `Qty | Rate | Amount` triple by arithmetic verification.
                last_dec = decimals[-1]
                second_last_dec = decimals[-2]
                last_val = last_dec[2]
                rate_val = second_last_dec[2]
                # Walk integers left of the rate column, nearest first.
                for entry in reversed(numerics):
                    if entry[0] >= second_last_dec[0]:
                        continue
                    if entry[3]:
                        continue  # integers only for qty candidate
                    qty_val = entry[2]
                    if qty_val <= 0 or last_val <= 0:
                        continue
                    expected = qty_val * rate_val
                    if abs(expected - last_val) / last_val <= 0.02:
                        rate_entry = second_last_dec
                        quantity_entry = entry
                        break

                if rate_entry is None:
                    # Arithmetic did not line up - fall back to right-most decimal.
                    rate_entry = last_dec
            elif len(decimals) == 1:
                rate_entry = decimals[0]
            else:
                # No decimals at all - rate stays the right-most numeric (legacy behaviour).
                rate_entry = numerics[-1]

            rate_pos = rate_entry[0]

            # ---- pick quantity if not already chosen by the Qty*Rate=Amount match above
            if quantity_entry is None:
                # Nearest integer to the LEFT of the rate.
                for entry in reversed(numerics):
                    if entry[0] >= rate_pos:
                        continue
                    if not entry[3]:  # integer-looking
                        quantity_entry = entry
                        break
                if quantity_entry is None:
                    # No integer to the left; fall back to any numeric to the left.
                    for entry in reversed(numerics):
                        if entry[0] < rate_pos:
                            quantity_entry = entry
                            break

            if quantity_entry is None:
                skipped_reasons.append(f"row {row_idx}: no quantity column to the left of rate")
                continue

            try:
                quantity = int(quantity_entry[2])
                rate = float(rate_entry[2])
            except (ValueError, TypeError):
                skipped_reasons.append(f"row {row_idx}: numeric conversion failed")
                continue

            if quantity <= 0 or rate <= 0:
                skipped_reasons.append(f"row {row_idx}: invalid qty/rate ({quantity}, {rate})")
                continue

            trades.append({
                "direction": direction,
                "quantity": quantity,
                "rate": rate,
                "source_line": joined,
            })

        if log:
            if trades:
                log.add("parse", "ok",
                        f"Tabular parser kept {len(trades)} trade(s) from {considered} BUY/SELL row(s)")
            else:
                log.add("parse", "warn",
                        f"Tabular parser found {considered} BUY/SELL row(s) but kept 0",
                        detail="\n".join(skipped_reasons[:20]))
        return trades

    # ------------------------------------------------------------------ legacy fallback parser

    def parse_trading_data(self, text):
        """Legacy line-based parser kept as a fallback when the tabular path finds nothing.

        Fixes vs original:
          * supports thousands-separated numbers (1,000.50)
          * skips obvious time/date tokens
          * positional rule: rate = right-most decimal numeric, quantity = integer numeric
            immediately to the left of the rate (avoids picking up an order ID).
        """
        if not text:
            return []

        trades = []
        lines = text.strip().split("\n")
        num_re = re.compile(r"-?\d{1,3}(?:,\d{3})+(?:\.\d+)?|-?\d+(?:\.\d+)?")

        for idx, raw_line in enumerate(lines):
            line = raw_line.strip()
            if not line:
                continue

            upper = line.upper()
            has_direction = "BUY" in upper or "SELL" in upper

            target_line = None
            direction = None
            if has_direction:
                target_line = line
                direction = "BUY" if "BUY" in upper else "SELL"
            elif idx > 0:
                prev_upper = lines[idx - 1].upper()
                if ("BUY" in prev_upper or "SELL" in prev_upper) and re.fullmatch(r"[\d\s,.]+", line):
                    target_line = line
                    direction = "BUY" if "BUY" in prev_upper else "SELL"

            if not target_line:
                continue

            tokens = target_line.split()
            numerics = []  # (pos, cleaned, value, has_decimal)
            for i, tok in enumerate(tokens):
                if _looks_like_time_or_date(tok):
                    continue
                m = num_re.fullmatch(tok.strip("()[]{}$₹€£%"))
                if not m:
                    continue
                cleaned = m.group(0).replace(",", "")
                try:
                    value = float(cleaned)
                except ValueError:
                    continue
                numerics.append((i, cleaned, value, "." in cleaned))

            if len(numerics) < 2:
                continue

            rate_entry = next((e for e in reversed(numerics) if e[3]), numerics[-1])
            rate_pos = rate_entry[0]
            quantity_entry = next((e for e in reversed(numerics) if e[0] < rate_pos and not e[3]), None)
            if quantity_entry is None:
                quantity_entry = next((e for e in reversed(numerics) if e[0] < rate_pos), None)
            if quantity_entry is None:
                continue

            try:
                quantity = int(quantity_entry[2])
                rate = float(rate_entry[2])
            except (ValueError, TypeError):
                continue

            if quantity > 0 and rate > 0:
                trades.append({
                    "direction": direction,
                    "quantity": quantity,
                    "rate": rate,
                    "source_line": target_line,
                })

        return trades

    # ------------------------------------------------------------------ orchestrator

    def extract_with_steps(self, image_path, echo=True):
        """Full pipeline for one image. Returns (trades, steps_list)."""
        log = StepLog(echo=echo, prefix=f"{Path(image_path).name}: ")
        log.add("start", "info", f"Begin processing {image_path}")

        raw_text, words = self.run_ocr(image_path, log=log)
        if not raw_text and not words:
            log.add("done", "error", "Stopped: OCR produced no text")
            return [], log.entries

        rows = self.build_table_rows(words, log=log)
        trades = self.parse_trades_from_rows(rows, log=log) if rows else []

        if not trades:
            log.add("parse", "info", "Falling back to line-based parser")
            trades = self.parse_trading_data(raw_text)
            if trades:
                log.add("parse", "ok", f"Fallback parser kept {len(trades)} trade(s)")
            else:
                log.add("parse", "warn", "Fallback parser also kept 0 trade(s)",
                        detail=raw_text[:1200])

        log.add("done", "ok" if trades else "warn",
                f"Finished with {len(trades)} trade(s)")
        return trades, log.entries

    def process_image(self, image_path):
        """Process a single image and extract trading data (uses the new step pipeline)."""
        print(f"Processing: {image_path}")
        trades, steps = self.extract_with_steps(image_path, echo=True)

        doc_name = Path(image_path).stem
        self.steps_by_doc[doc_name] = steps

        if not trades:
            return 0

        self.documents[doc_name] = trades
        self.trading_data.extend(trades)
        return len(trades)
    
    def process_images(self, image_paths):
        """Process multiple images"""
        print("=" * 60)
        print("TRADING DATA EXTRACTION")
        print("=" * 60)
        
        total_trades = 0
        for image_path in image_paths:
            if isinstance(image_path, str):
                image_path = Path(image_path)
            
            if not image_path.exists():
                print(f"⚠ File not found: {image_path}")
                print(f"  (Check that spaces and special characters are correct)")
                continue
            
            if not image_path.is_file():
                print(f"⚠ Not a file: {image_path}")
                continue
            
            trades = self.process_image(image_path)
            total_trades += trades
        
        print(f"\n✓ Total trades extracted: {total_trades}")
        return total_trades > 0
    
    def calculate_values(self):
        """Calculate (Quantity/250) * Rate for all trades"""
        results = {
            'buy': [],
            'sell': []
        }
        
        for trade in self.trading_data:
            calculated = (trade['quantity'] / 250) * trade['rate']
            trade['calculated'] = calculated
            
            if trade['direction'] == 'BUY':
                results['buy'].append(trade)
            else:
                results['sell'].append(trade)
        
        return results
    
    def generate_excel(self, output_path='trading_analysis_output.xlsx'):
        """Generate Excel file with analysis"""
        print("\n" + "=" * 60)
        print("GENERATING EXCEL REPORT")
        print("=" * 60)
        
        # Calculate values
        results = self.calculate_values()
        
        buy_total = sum(t['calculated'] for t in results['buy'])
        sell_total = sum(t['calculated'] for t in results['sell'])
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Create summary sheet first
        summary_ws = wb.active
        summary_ws.title = "Summary"
        
        self._create_summary_sheet(summary_ws, results, buy_total, sell_total)
        
        # Create detail sheets for each document
        for doc_name, trades in self.documents.items():
            if trades:
                ws = wb.create_sheet(doc_name[:31])  # Excel sheet name limit is 31 chars
                self._create_detail_sheet(ws, trades)
        
        # Save
        wb.save(output_path)
        print(f"\n✓ Excel report generated: {output_path}")
        
        # Print summary
        print("\n" + "=" * 60)
        print("ANALYSIS SUMMARY")
        print("=" * 60)
        print(f"BUY Total:  {buy_total:,.4f}")
        print(f"SELL Total: {sell_total:,.4f}")
        print(f"NET (BUY - SELL): {buy_total - sell_total:,.4f}")
        print("=" * 60)
        
        return output_path
    
    def _create_summary_sheet(self, ws, results, buy_total, sell_total):
        """Create summary sheet"""
        ws['A1'] = "TRADING ANALYSIS SUMMARY"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Overall summary
        ws['A3'] = "OVERALL TOTALS"
        ws['A3'].font = Font(bold=True, size=12)
        
        ws['A4'] = "Type"
        ws['B4'] = "Total Value"
        ws['C4'] = "Count"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}4']
            cell.fill = header_fill
            cell.font = header_font
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        # BUY row
        ws['A5'] = "BUY"
        ws['B5'] = buy_total
        ws['C5'] = len(results['buy'])
        ws['B5'].number_format = '0.0000'
        ws['B5'].font = Font(color="008000", bold=True)
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}5'].border = thin_border
        
        # SELL row
        ws['A6'] = "SELL"
        ws['B6'] = sell_total
        ws['C6'] = len(results['sell'])
        ws['B6'].number_format = '0.0000'
        ws['B6'].font = Font(color="FF0000", bold=True)
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}6'].border = thin_border
        
        # NET row
        ws['A7'] = "NET (BUY - SELL)"
        ws['B7'] = buy_total - sell_total
        ws['B7'].number_format = '0.0000'
        ws['B7'].font = Font(bold=True)
        
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}7']
            cell.border = thin_border
            if col in ['A', 'B']:
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
    
    def _create_detail_sheet(self, ws, trades):
        """Create detail sheet for a document"""
        # Headers
        headers = ["Sell/Buy", "Quantity", "Market Rate", "Calculated Value (Qty/250 × Rate)"]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        buy_total = 0
        sell_total = 0
        
        # Add data
        for trade in trades:
            ws.append([
                trade['direction'],
                trade['quantity'],
                trade['rate'],
                trade['calculated']
            ])
            
            if trade['direction'] == 'BUY':
                buy_total += trade['calculated']
            else:
                sell_total += trade['calculated']
        
        # Format data
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = thin_border
                if cell.column in [3, 4]:
                    cell.number_format = '0.0000'
        
        # Add summary
        summary_row = ws.max_row + 2
        ws[f'A{summary_row}'] = "DOCUMENT SUMMARY"
        ws[f'A{summary_row}'].font = Font(bold=True)
        
        summary_row += 1
        ws[f'A{summary_row}'] = "Type"
        ws[f'B{summary_row}'] = "Total"
        
        for cell in [ws[f'A{summary_row}'], ws[f'B{summary_row}']]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.border = thin_border
        
        summary_row += 1
        ws[f'A{summary_row}'] = "BUY"
        ws[f'B{summary_row}'] = buy_total
        ws[f'B{summary_row}'].number_format = '0.0000'
        ws[f'B{summary_row}'].font = Font(bold=True, color="008000")
        
        for cell in [ws[f'A{summary_row}'], ws[f'B{summary_row}']]:
            cell.border = thin_border
        
        summary_row += 1
        ws[f'A{summary_row}'] = "SELL"
        ws[f'B{summary_row}'] = sell_total
        ws[f'B{summary_row}'].number_format = '0.0000'
        ws[f'B{summary_row}'].font = Font(bold=True, color="FF0000")
        
        for cell in [ws[f'A{summary_row}'], ws[f'B{summary_row}']]:
            cell.border = thin_border
        
        summary_row += 1
        ws[f'A{summary_row}'] = "NET (BUY - SELL)"
        ws[f'B{summary_row}'] = buy_total - sell_total
        ws[f'B{summary_row}'].number_format = '0.0000'
        ws[f'B{summary_row}'].font = Font(bold=True)
        
        for cell in [ws[f'A{summary_row}'], ws[f'B{summary_row}']]:
            cell.border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30


def main():
    """Main function"""
    print("\n")
    print("╔" + "═" * 58 + "╗")
    print("║" + " " * 58 + "║")
    print("║" + "  TRADING DATA ANALYSIS TOOL - OCR VERSION".center(58) + "║")
    print("║" + " " * 58 + "║")
    print("╚" + "═" * 58 + "╝")
    
    # Get image paths from user
    print("\nEnter image file paths (one per line, press Enter twice to finish):")
    print("(Supported formats: JPG, PNG, JPEG, TIFF)")
    print("(Note: Paths with spaces and special characters are supported)")
    print()
    
    SUPPORTED_FORMATS = {'.jpg', '.jpeg', '.png', '.tiff', '.tif'}
    image_paths = []
    while True:
        path = input(f"Image {len(image_paths) + 1}: ").strip()
        if not path:
            if image_paths:
                break
            else:
                print("Please enter at least one image path.")
                continue
        
        # Remove quotes if present (handles copy-paste from Windows Explorer)
        path = path.strip('"').strip("'")
        
        # Validate file exists
        if not os.path.exists(path):
            print(f"  ⚠ File not found: {path}")
            print(f"    Please verify the path is correct (including spaces and parentheses)")
            continue
        
        # Validate file format
        file_ext = Path(path).suffix.lower()
        if file_ext not in SUPPORTED_FORMATS:
            print(f"  ⚠ Unsupported format: {file_ext}")
            print(f"    Supported formats: {', '.join(SUPPORTED_FORMATS)}")
            continue
        
        image_paths.append(path)
        print(f"  ✓ Added: {Path(path).name}")
    
    if not image_paths:
        print("No images provided. Exiting.")
        return
    
    # Create extractor
    extractor = TradingDataExtractor()
    
    # Process images
    if not extractor.process_images(image_paths):
        print("Failed to extract trading data from images.")
        return
    
    # Generate output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"trading_analysis_{timestamp}.xlsx"
    
    # Generate Excel
    output_path = extractor.generate_excel(output_file)
    
    print(f"\n✓ Analysis complete! File saved as: {output_file}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Error: {e}")
        sys.exit(1)
